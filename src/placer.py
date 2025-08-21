import csv
import io
import random
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple, Set, FrozenSet
from collections import Counter


# ---------- Data classes ----------
@dataclass(frozen=True)
class DreTileType:
    name: str
    length: int
    markers: List[int]        # DRE activation markers (1-based)
    duo: FrozenSet[int]       # duo markers (1-based)

@dataclass(frozen=True)
class HDTileType:
    name: str
    length: int
    pink: List[int]           # 1-based
    grey: List[int]           # 1-based
    duo: FrozenSet[int]       # duo markers (1-based)

@dataclass
class PlacedTile:
    name: str
    start: int  # 1-based top row
    length: int

@dataclass
class Grid:
    rows: int
    dre_col: List[Optional[str]] = field(init=False)
    hda_col: List[Optional[str]] = field(init=False)
    hdb_col: List[Optional[str]] = field(init=False)
    dre_tiles: List[PlacedTile] = field(default_factory=list)
    hda_tiles: List[PlacedTile] = field(default_factory=list)
    hdb_tiles: List[PlacedTile] = field(default_factory=list)
    activated_rows_available: Set[int] = field(default_factory=set)
    activated_rows_all: Set[int] = field(default_factory=set)
    dre_duo_signatures: Set[FrozenSet[int]] = field(default_factory=set)  # duo sets present
    dre_duo_anchors: Dict[FrozenSet[int], List[int]] = field(default_factory=dict)  # duo -> [DRE starts]
    dre_duo_available: Dict[FrozenSet[int], Set[int]] = field(default_factory=dict) # duo -> available starts

    def __post_init__(self):
        self.dre_col = [None] * self.rows
        self.hda_col = [None] * self.rows
        self.hdb_col = [None] * self.rows

    def free_span(self, col: str) -> List[Tuple[int,int]]:
        arr = {"dre": self.dre_col, "hda": self.hda_col, "hdb": self.hdb_col}[col]
        spans, i = [], 0
        while i < self.rows:
            if arr[i] is None:
                j = i
                while j < self.rows and arr[j] is None:
                    j += 1
                spans.append((i+1, j - i))  # 1-based start, length
                i = j
            else:
                i += 1
        return spans

    def place_tile(self, col: str, tile_name: str, start_row: int, length: int) -> bool:
        arr = {"dre": self.dre_col, "hda": self.hda_col, "hdb": self.hdb_col}[col]
        idx0 = start_row - 1
        if idx0 < 0 or idx0 + length > self.rows:
            return False
        if any(arr[idx0:idx0+length]):
            return False
        for k in range(idx0, idx0+length):
            arr[k] = tile_name
        pt = PlacedTile(tile_name, start_row, length)
        (self.dre_tiles if col=="dre" else self.hda_tiles if col=="hda" else self.hdb_tiles).append(pt)
        return True

    def compute_dre_activation(self, dre_types: Dict[str,'DreTileType']):
        self.activated_rows_all.clear()
        self.dre_duo_signatures.clear()
        self.dre_duo_anchors.clear()
        self.dre_duo_available.clear()
        for pt in self.dre_tiles:
            tt = dre_types[pt.name]
            for m in tt.markers:
                row = pt.start + m - 1
                if 1 <= row <= self.rows:
                    self.activated_rows_all.add(row)
            if tt.duo:
                self.dre_duo_signatures.add(tt.duo)
                self.dre_duo_anchors.setdefault(tt.duo, []).append(pt.start)
        self.activated_rows_available = set(sorted(self.activated_rows_all))
        for duo, starts in self.dre_duo_anchors.items():
            self.dre_duo_available[duo] = set(starts)

    def allow_duo_for_hd(self, t: 'HDTileType') -> bool:
        # Exact duo set required in grid
        return (not t.duo) or (t.duo in self.dre_duo_signatures)

    def try_place_hd_matching(self, which_col: str, t: 'HDTileType') -> Optional[int]:
        if not self.allow_duo_for_hd(t): return None
        spans = self.free_span(which_col)
        random.shuffle(spans)

        # Pink/grey adjacency mapping
        adjacency = {}
        for p in t.pink:
            opts = {p}
            for g in t.grey:
                if abs(g - p) == 1: opts.add(g)
            adjacency[p] = sorted(opts)

        # Candidate start rows: if duo, must align to available DRE anchor with same duo
        if t.duo:
            candidate_starts = sorted(self.dre_duo_available.get(t.duo, []))
        else:
            candidate_starts = []
            for start, avail_len in spans:
                for sr in range(start, start + avail_len - t.length + 1):
                    candidate_starts.append(sr)

        for start, avail_len in spans:
            span_min = start
            span_max = start + avail_len - t.length
            for sr in candidate_starts:
                if sr < span_min or sr > span_max: continue
                # Check pink/grey against activated rows
                needed, ok = [], True
                for p, opts in adjacency.items():
                    cand_rows = [sr + o - 1 for o in opts]
                    hit = next((cr for cr in cand_rows if cr in self.activated_rows_available), None)
                    if hit is None: ok = False; break
                    needed.append(hit)
                if ok and self.place_tile(which_col, t.name, sr, t.length):
                    for r in needed: self.activated_rows_available.discard(r)
                    if t.duo and sr in self.dre_duo_available.get(t.duo, set()):
                        self.dre_duo_available[t.duo].discard(sr)  # consume the anchor
                    return sr
        return None

    def place_hd_anywhere(self, which_col: str, t: 'HDTileType') -> Optional[int]:
        if not self.allow_duo_for_hd(t): return None
        spans = self.free_span(which_col)
        if t.duo:
            cands = sorted(self.dre_duo_available.get(t.duo, []))
            for start, avail_len in spans:
                span_min, span_max = start, start + avail_len - t.length
                for sr in cands:
                    if span_min <= sr <= span_max:
                        if self.place_tile(which_col, t.name, sr, t.length):
                            self.dre_duo_available[t.duo].discard(sr)
                            return sr
        else:
            for start, avail_len in spans:
                if avail_len >= t.length and self.place_tile(which_col, t.name, start, t.length):
                    return start
        return None

# ---------- CSV helpers ----------
def parse_int_list(s: str) -> List[int]:
    if not s: return []
    s = s.strip().replace(",", " ").replace(";", " ")
    out = []
    for tok in s.split():
        try:
            n = int(float(tok))
            if n > 0: out.append(n)
        except: pass
    return out

def _csv_reader_from_bytes(b: bytes):
    # filter out comment lines starting with //
    it = io.StringIO(b.decode("utf-8", errors="replace"))
    def _lines():
        for line in it:
            if line.lstrip().startswith("//"): continue
            yield line
    return csv.DictReader(_lines())

def load_tiles_from_bytes(tiles_csv_bytes: bytes):
    dre_types: Dict[str,DreTileType] = {}
    hd_types: Dict[str,HDTileType] = {}
    reader = _csv_reader_from_bytes(tiles_csv_bytes)
    headers = {h.lower(): h for h in (reader.fieldnames or [])}
    req = {"name","group","length"}
    if not req.issubset({h.lower() for h in headers}):
        raise ValueError("tile_characteristics.csv must include name, group, length")
    for row in reader:
        name = row.get(headers.get("name")) or row.get("name")
        group = (row.get(headers.get("group")) or "").strip().upper()
        length_str = row.get(headers.get("length")) or ""
        if not name or not group or not length_str: continue
        length = int(float(length_str))
        dre_m = parse_int_list(row.get(headers.get("dre_markers")) if headers.get("dre_markers") else row.get("dre_markers",""))
        pink_m = parse_int_list(row.get(headers.get("pink_markers")) if headers.get("pink_markers") else row.get("pink_markers",""))
        grey_m = parse_int_list(row.get(headers.get("grey_markers")) if headers.get("grey_markers") else row.get("grey_markers",""))
        duo_m  = parse_int_list(row.get(headers.get("duo_markers")) if headers.get("duo_markers") else row.get("duo_markers",""))
        duo_sig = frozenset(duo_m)
        if group == "DRE":
            dre_types[name] = DreTileType(name, length, dre_m, duo_sig)
        elif group == "HD":
            hd_types[name] = HDTileType(name, length, pink_m, grey_m, duo_sig)
        else:
            raise ValueError(f"Unknown group '{group}' in tile_characteristics.csv")
    return dre_types, hd_types

def load_requirements_from_bytes(reqs_csv_bytes: bytes) -> Dict[str,int]:
    reqs: Dict[str,int] = {}
    reader = _csv_reader_from_bytes(reqs_csv_bytes)
    headers = {h.lower(): h for h in (reader.fieldnames or [])}
    need = {"name","count"}
    if not need.issubset({h.lower() for h in headers}):
        raise ValueError("tile_requirements.csv must include name,count")
    for row in reader:
        name = row.get(headers.get("name")) or row.get("name")
        cnt  = row.get(headers.get("count")) or row.get("count")
        if not name: continue
        reqs[name] = int(float(cnt or 0))
    return reqs

# ---------- Placement ----------
def more_space_col(g: Grid) -> str:
    a = sum(1 for x in g.hda_col if x is None)
    b = sum(1 for x in g.hdb_col if x is None)
    return "hda" if a >= b else "hdb"

def place_all(dre_types, hd_types, reqs, grids_count: int, rows: int, seed: int):
    random.seed(seed)
    grids = [Grid(rows=rows) for _ in range(grids_count)]

    dre_pool, hd_pool, missing = [], [], []
    for name, cnt in reqs.items():
        if name in dre_types: dre_pool.extend([name]*cnt)
        elif name in hd_types: hd_pool.extend([name]*cnt)
        else: missing.append(name)

    # Enforce 1:1 EXCDRE↔EXCHD (your explicit parity rule)
    if "EXCDRE" in reqs and "EXCHD" in reqs and reqs["EXCDRE"] != reqs["EXCHD"]:
        raise ValueError(f"Requirement mismatch: EXCDRE ({reqs['EXCDRE']}) and EXCHD ({reqs['EXCHD']}) must be equal.")

    # Place DRE longest-first, round-robin
    dre_pool.sort(key=lambda n: dre_types[n].length, reverse=True)
    gi = 0
    for name in dre_pool:
        placed = False
        for _ in range(grids_count):
            g = grids[gi]
            for start, avail_len in g.free_span("dre"):
                if avail_len >= dre_types[name].length:
                    g.place_tile("dre", name, start, dre_types[name].length)
                    placed = True; break
            if placed: break
            gi = (gi + 1) % grids_count
        if not placed:
            for j in range(grids_count):
                g = grids[j]
                for start, avail_len in g.free_span("dre"):
                    if avail_len >= dre_types[name].length:
                        g.place_tile("dre", name, start, dre_types[name].length)
                        placed = True; break
                if placed: break

    for g in grids: g.compute_dre_activation(dre_types)

    def needs_matching(n: str) -> bool:
        return len(hd_types[n].pink) > 0

    hd_with = [n for n in hd_pool if needs_matching(n)]
    hd_without = [n for n in hd_pool if not needs_matching(n)]
    hd_with.sort(key=lambda n: (len(hd_types[n].pink), hd_types[n].length), reverse=True)

    unplaced: List[str] = []

    for name in hd_with:
        t = hd_types[name]
        order = list(range(grids_count))
        random.shuffle(order)
        order.sort(key=lambda idx: len(grids[idx].activated_rows_available), reverse=True)
        done = False
        for idx in order:
            g = grids[idx]
            if len(g.activated_rows_available) < len(t.pink): continue
            for which in [more_space_col(g), "hda" if more_space_col(g)=="hdb" else "hdb"]:
                if g.try_place_hd_matching(which, t) is not None:
                    done = True; break
            if done: break
        if not done: unplaced.append(name)

    for name in hd_without:
        t = hd_types[name]
        done = False
        order = list(range(grids_count))
        random.shuffle(order)
        for idx in order:
            g = grids[idx]
            for which in ["hda", "hdb"]:
                if g.place_hd_anywhere(which, t) is not None:
                    done = True; break
            if done: break
        if not done: unplaced.append(name)

    return grids, unplaced, missing

# ---------- Excel export (to bytes) ----------
def export_workbook_to_bytes(grids, rows, dre_types, hd_types, reqs, unplaced, missing) -> bytes:
    # Import here to avoid heavy import cost at module import time
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side, Font
    from openpyxl.utils import get_column_letter

    GRID_COUNT = len(grids)
    cols = []
    for i in range(1, GRID_COUNT+1):
        cols += [f"Dre{i}", f"HDa{i}", f"HDb{i}"]

    table_data = []
    for r in range(1, rows+1):
        row_vals = []
        for g in grids:
            row_vals.extend([
                g.dre_col[r-1] or "",
                g.hda_col[r-1] or "",
                g.hdb_col[r-1] or "",
            ])
        table_data.append(row_vals)

    holder_counts = {}
    for i, g in enumerate(grids, start=1):
        holder_counts[f"Dre{i}"] = len(g.dre_tiles)
        holder_counts[f"HDa{i}"] = len(g.hda_tiles)
        holder_counts[f"HDb{i}"] = len(g.hdb_tiles)

    GREEN  = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    PINK   = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    ORANGE = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    DUO    = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    COUNT_FILL  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    THIN = Side(style="thin", color="000000")
    THIN_BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)

    wb = Workbook()
    ws = wb.active; ws.title = "Grids"

    ws.cell(row=1, column=1, value="Row").fill = HEADER_FILL
    ws.cell(row=1, column=1).font = Font(bold=True)
    for j, h in enumerate(cols, start=2):
        ws.cell(row=1, column=j, value=h).fill = HEADER_FILL
        ws.cell(row=1, column=j).font = Font(bold=True)

    for r in range(rows):
        ws.cell(row=r+2, column=1, value=r+1)
        for j, val in enumerate(table_data[r], start=2):
            ws.cell(row=r+2, column=j, value=val)

    counts_excel_row = rows + 2
    ws.cell(row=counts_excel_row, column=1, value="TileCount").fill = COUNT_FILL
    ws.cell(row=counts_excel_row, column=1).font = Font(bold=True)
    for j, h in enumerate(cols, start=2):
        ws.cell(row=counts_excel_row, column=j, value=holder_counts[h]).fill = COUNT_FILL
        ws.cell(row=counts_excel_row, column=j).font = Font(bold=True)

    max_col = 1 + len(cols)
    for r in range(1, counts_excel_row+1):
        for c in range(1, max_col+1):
            ws.cell(row=r, column=c).border = THIN_BORDER
    for c in range(1, max_col+1):
        ws.column_dimensions[get_column_letter(c)].width = 16 if c != 1 else 8

    def holder_col_index(grid_idx: int, holder: str) -> int:
        base = (grid_idx-1)*3 + 2
        return base if holder=="Dre" else (base+1 if holder=="HDa" else base+2)

    for gi, g in enumerate(grids, start=1):
        # DRE markers
        for pt in g.dre_tiles:
            tt = dre_types[pt.name]
            col_idx = holder_col_index(gi, "Dre")
            for m in tt.markers:
                r_abs = pt.start + m - 1
                if 1 <= r_abs <= rows:
                    ws.cell(row=1 + r_abs, column=col_idx).fill = GREEN
            for dm in tt.duo:
                r_abs = pt.start + dm - 1
                if 1 <= r_abs <= rows:
                    ws.cell(row=1 + r_abs, column=col_idx).fill = DUO

        # HDa markers
        for pt in g.hda_tiles:
            if pt.name in hd_types:
                t = hd_types[pt.name]; col_idx = holder_col_index(gi, "HDa")
                for p in t.pink:
                    r_abs = pt.start + p - 1
                    if 1 <= r_abs <= rows:
                        ws.cell(row=1 + r_abs, column=col_idx).fill = PINK
                for gr in t.grey:
                    r_abs = pt.start + gr - 1
                    if 1 <= r_abs <= rows:
                        ws.cell(row=1 + r_abs, column=col_idx).fill = ORANGE
                for dm in t.duo:
                    r_abs = pt.start + dm - 1
                    if 1 <= r_abs <= rows:
                        ws.cell(row=1 + r_abs, column=col_idx).fill = DUO

        # HDb markers
        for pt in g.hdb_tiles:
            if pt.name in hd_types:
                t = hd_types[pt.name]; col_idx = holder_col_index(gi, "HDb")
                for p in t.pink:
                    r_abs = pt.start + p - 1
                    if 1 <= r_abs <= rows:
                        ws.cell(row=1 + r_abs, column=col_idx).fill = PINK
                for gr in t.grey:
                    r_abs = pt.start + gr - 1
                    if 1 <= r_abs <= rows:
                        ws.cell(row=1 + r_abs, column=col_idx).fill = ORANGE
                for dm in t.duo:
                    r_abs = pt.start + dm - 1
                    if 1 <= r_abs <= rows:
                        ws.cell(row=1 + r_abs, column=col_idx).fill = DUO

    ws.freeze_panes = "B2"

    # Summary
    placed_counter = Counter()
    for g in grids:
        for pt in g.dre_tiles + g.hda_tiles + g.hdb_tiles:
            placed_counter[pt.name] += 1

    total_dre = sum(placed_counter.get(name, 0) for name in dre_types.keys())
    total_hd  = sum(placed_counter.get(name, 0) for name in hd_types.keys())

    sum_ws = wb.create_sheet("Summary")
    sum_ws["A1"] = "Summary"; sum_ws["A1"].font = Font(bold=True, size=14)
    sum_ws["A3"] = "Total Dre tiles"; sum_ws["B3"] = total_dre
    sum_ws["A4"] = "Total HD tiles";  sum_ws["B4"] = total_hd

    r = 6
    sum_ws[f"A{r}"] = "Per-type counts"; sum_ws[f"A{r}"].font = Font(bold=True); r += 1
    for name in sorted(dre_types.keys()):
        sum_ws[f"A{r}"] = name; sum_ws[f"B{r}"] = placed_counter.get(name, 0); r += 1
    for name in sorted(hd_types.keys()):
        sum_ws[f"A{r}"] = name; sum_ws[f"B{r}"] = placed_counter.get(name, 0); r += 1

    r += 1
    sum_ws[f"A{r}"] = "Tiles requested but not placed"; sum_ws[f"A{r}"].font = Font(bold=True); r += 1
    unplaced_counts = Counter(unplaced)
    if unplaced_counts:
        for name, cnt in sorted(unplaced_counts.items()):
            sum_ws[f"A{r}"] = name; sum_ws[f"B{r}"] = cnt; r += 1
    else:
        sum_ws[f"A{r}"] = "(none)"; r += 1

    r += 1
    sum_ws[f"A{r}"] = "Names in requirements but missing from characteristics"; sum_ws[f"A{r}"].font = Font(bold=True); r += 1
    if missing:
        for name in sorted(missing):
            sum_ws[f"A{r}"] = name; r += 1
    else:
        sum_ws[f"A{r}"] = "(none)"; r += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# ---------- Public entry ----------
def place_and_export(tiles_csv_bytes: bytes, reqs_csv_bytes: bytes, grids: int, rows: int, seed: int) -> bytes:
    dre_types, hd_types = load_tiles_from_bytes(tiles_csv_bytes)
    reqs = load_requirements_from_bytes(reqs_csv_bytes)
    grids_obj, unplaced, missing = place_all(dre_types, hd_types, reqs, grids, rows, seed)
    return export_workbook_to_bytes(grids_obj, rows, dre_types, hd_types, reqs, unplaced, missing)
